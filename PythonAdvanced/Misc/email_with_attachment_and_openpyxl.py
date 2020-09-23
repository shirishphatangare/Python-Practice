
import argparse
import csv
import requests
import logging
import errno
import datetime
from dateutil.relativedelta import relativedelta # pip install python-dateutil
import smtplib
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.chart import BarChart, Reference
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def valid_date(s):
    """
        Objective: This function validates date in command line arguments
        Input: string representing date and time
        Output : datetime object
        Use it: type=valid_date
    """
    try:
        # strptime() method creates a datetime object from a given string (representing date and time).
        return datetime.datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        msg = "Not a valid date: '{0}'.".format(s)
        raise argparse.ArgumentTypeError(msg)

parser = argparse.ArgumentParser(description='Generate bandwidth utilization report. fromdate and todate arguments are optional. If not given by default report will be generated from 1st day of previous month to last day of previous month.')
parser.add_argument('-f', '--fromdate', type=valid_date, help="The From Date - format YYYY-MM-DD. Default is start_of_last_month")
parser.add_argument('-t', '--todate', type=valid_date, help="The To Date - format YYYY-MM-DD. Default is end_of_last_month")
args = parser.parse_args()

#If args are not provided, Report will be generated from 1st day of previous month to last day of previous month
if args.fromdate:
    from_date = args.fromdate
    from_date_report_name = from_date
    print(f"{from_date}")
else:
    from_date = "start_of_last_month"
    today = datetime.date.today()
    first = today.replace(day=1)
    from_date_report_name = (first - relativedelta(months=1)).strftime("%Y-%m-%d")

if args.todate:
    to_date = args.todate
    to_date_report_name = to_date
    print(f"{to_date}")
else:
    to_date = "end_of_last_month"
    today = datetime.date.today()
    first = today.replace(day=1)
    to_date_report_name = (first - datetime.timedelta(days=1)).strftime("%Y-%m-%d")

input_file = "resources/input.csv"
output_file = f"resources/Bandwidth_Utilization_Report_{from_date_report_name}_{to_date_report_name}.xlsx"

# http://server-name/api/v2.1/cdt_port/?fields=cdt_device.name,name,id,ifSpeed,ifTitle,TxUtil,RxUtil&RxUtil_formats=max,avg,min&TxUtil_formats=max,avg,min&RxUtil_timefilter=range=start_of_today -30d to start_of_today&TxUtil_timefilter=range=start_of_today -30d to start_of_today&cdt_device.name_filter=IS('MEM25FW-FXF-F2AR27-r02')&name_filter=IS('Te1/3/0')&links=none
def create_staseeker_url(device, interface):
    """
        Objective: This function sends request to staseeker REST endpoint
        Input: device and interface names
        Output : staseeker_url
        Use it: create_staseeker_url(device, interface)
    """
    # API root endpoint
    #rootUri = "server-name"

    # specify target endpoint
    targetUri = "cdt_port/"

    # specify fields to be returned and filters to use
    getFields = "?fields=cdt_device.name,name,id,ifSpeed,ifTitle,TxUtil,RxUtil"

    # specify data formats for each metric
    dataFormats = "&RxUtil_formats=max,avg,min&TxUtil_formats=max,avg,min"

    # specify time filters, one for each timeseries metric being requested
    timeFilters = f"&RxUtil_timefilter=range={from_date} to {to_date}&TxUtil_timefilter=range={from_date} to {to_date}"

    # optional response formatting
    rspFormat = "&links=none"

    # specify data filters to select only those devices\interfaces you want to report on
    dataFilters = f"&cdt_device.name_filter=IS('{device}')&name_filter=IS('{interface}')"

    # set URL to be used
    staseeker_url =  rootUri + targetUri + getFields + dataFormats + timeFilters + dataFilters + rspFormat
    return staseeker_url;

def send_statseeker_request(staseeker_url):
    """
        Objective: This function sends request to staseeker REST endpoint
        Input: staseeker endpoint
        Output : staseeker_response
        Use it: send_statseeker_request(staseeker_url)
    """
    # headers
    headers = {"Accept": "application/json", "Content-Type": "application/json"}

    # credentials
    user = "user"
    pword = "password"

    # send request
    staseeker_response = requests.get(staseeker_url, headers=headers, auth=(user, pword))
    return staseeker_response;

def parse_and_save_data(response, opco):
    """
        Objective: This function parse and save data in utilization_data_worksheet
        Input: Staseeker response
        Output : None
        Use it: parse_and_save_data(response, opco)
    """
    # Assigning value to global variable in local scope treats it as a local variable. That is why use global keyword
    global report_rows, chart_rows

    if len(response.json()['data']['objects'])  > 0 and len(response.json()['data']['objects'][0]['data']) > 0:
        utilization_details = response.json()['data']['objects'][0]['data'][0]
        print(utilization_details)
        print(20 * "-")

        # Update data worksheet
        utilization_data_worksheet.cell(row=report_rows, column=1).value = opco
        utilization_data_worksheet.cell(row=report_rows, column=2).value = utilization_details['cdt_device.name']
        utilization_data_worksheet.cell(row=report_rows, column=3).value = utilization_details['name']
        utilization_data_worksheet.cell(row=report_rows, column=4).value = utilization_details['id']
        utilization_data_worksheet.cell(row=report_rows, column=5).value = utilization_details['ifSpeed'] / (10 ** 9)
        utilization_data_worksheet.cell(row=report_rows, column=6).value = utilization_details['ifTitle']
        utilization_data_worksheet.cell(row=report_rows, column=7).value = utilization_details['RxUtil']['min']
        utilization_data_worksheet.cell(row=report_rows, column=8).value = utilization_details['RxUtil']['avg']
        utilization_data_worksheet.cell(row=report_rows, column=9).value = utilization_details['RxUtil']['max']
        utilization_data_worksheet.cell(row=report_rows, column=10).value = utilization_details['TxUtil']['min']
        utilization_data_worksheet.cell(row=report_rows, column=11).value = utilization_details['TxUtil']['avg']
        utilization_data_worksheet.cell(row=report_rows, column=12).value = utilization_details['TxUtil']['max']

        # Update Chart worksheet
        utilization_chart_worksheet.cell(row=chart_rows, column=1).value = utilization_details['cdt_device.name']
        utilization_chart_worksheet.cell(row=chart_rows, column=2).value = utilization_details['name']
        utilization_chart_worksheet.cell(row=chart_rows, column=3).value = utilization_details['RxUtil']['avg']
        utilization_chart_worksheet.cell(row=chart_rows, column=4).value = utilization_details['RxUtil']['max']
        utilization_chart_worksheet.cell(row=chart_rows, column=5).value = utilization_details['TxUtil']['avg']
        utilization_chart_worksheet.cell(row=chart_rows, column=6).value = utilization_details['TxUtil']['max']

        report_rows += 1
        chart_rows += 1

        report_workbook.save(output_file)

def prepare_utilization_data_worksheet():
    """
        Objective: This function prepares utilization data worksheet in report
        Input: None
        Output : None
        Use it: prepare_utilization_data_worksheet()
    """
    try:
        input_fh = open(input_file, 'rt')
        # With DictReader results are interpreted as a dictionary, where the header row is the key, and other rows are values.
        reader = csv.DictReader(input_fh)

        for device_and_interface in reader:
            #print(device_and_interface)
            device = device_and_interface.get('Device Name')
            interface = device_and_interface.get('Interface Name')
            staseeker_url = create_staseeker_url(device, interface)

            print(f"Sending request for {device} and {interface}")
            staseeker_response = send_statseeker_request(staseeker_url)
            opco = device_and_interface.get('OpCo')
            parse_and_save_data(staseeker_response, opco)
    except OSError as e:
        if e.errno == errno.ENOENT:
            logging.error('File not found')
        elif e.errno == errno.EACCES:
            logging.error('Permission denied')
        else:
            logging.error('Unexpected error: % d', e.errno)
    finally:
        input_fh.close()

def prepare_utilization_chart_worksheet():
    """
        Objective: This function prepares a Bar chart in utilization chart worksheet
        Input: None
        Output : None
        Use it: prepare_utilization_chart_worksheet()
    """
    # inbound chart
    inbound_chart = BarChart() # Bar chart
    inbound_chart.height = 15  # default is 7.5
    inbound_chart.width = 30  # default is 15
    #inbound_chart.grouping = "stacked"
    #inbound_chart.overlap = 100
    #inbound_chart.type = "col"
    #inbound_chart.style = 13
    inbound_chart.title = "Bandwidth Utilization Chart - Inbound - Avg/Max (util %)"
    inbound_chart.y_axis.title = 'Avg/Max Utilization(%)'
    inbound_chart.x_axis.title = 'Device'

    inbound_data = Reference(worksheet=utilization_chart_worksheet,
                     min_row=1, min_col=3,
                     max_row=chart_rows - 1, max_col=4)

    inbound_chart.add_data(inbound_data, titles_from_data=True)

    categories = Reference(worksheet=utilization_chart_worksheet,
                           min_row=2, min_col=1,
                           max_row=chart_rows - 1, max_col=1)

    inbound_chart.set_categories(categories)
    utilization_chart_worksheet.add_chart(inbound_chart, "H2")

    #outbound chart
    outbound_chart = BarChart() # Bar chart
    outbound_chart.height = 15  # default is 7.5
    outbound_chart.width = 30  # default is 15
    #outbound_chart.grouping = "stacked"
    #outbound_chart.overlap = 100
    #outbound_chart.type = "col"
    #outbound_chart.style = 13
    outbound_chart.title = "Bandwidth Utilization Chart - Outbound - Avg/Max (util %)"
    outbound_chart.y_axis.title = 'Avg/Max Utilization(%)'
    outbound_chart.x_axis.title = 'Device'

    outbound_data = Reference(worksheet=utilization_chart_worksheet,
                     min_row=1, min_col=5,
                     max_row=chart_rows - 1, max_col=6)

    outbound_chart.add_data(outbound_data, titles_from_data=True)

    categories = Reference(worksheet=utilization_chart_worksheet,
                           min_row=2, min_col=1,
                           max_row=chart_rows - 1, max_col=1)

    outbound_chart.set_categories(categories)
    utilization_chart_worksheet.add_chart(outbound_chart, "H32")


    report_workbook.save(output_file)


def send_email_with_worksheet_attachment():
    """
        Objective: This function sends an email to the distribution list with utilization report worksheet as an attachment
        Input: None
        Output : None
        Use it: send_email_with_worksheet_attachment()
    """
    print(f"Sending email..")
    filename = output_file.split('/')[1]

    subject = f"Monthly Report - {filename}"
    body = f"This is an email with monthly report - {filename} as an attachment"
    sender_email = "sender_email"
    receiver_email = "receiver_email"
    smtp_server = "smtp_server"
    smtp_port = "smtp_port"

    # Create a multipart message and set headers
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject

    # Add body to email
    message.attach(MIMEText(body, "plain"))

    # Open PDF file in binary mode
    with open(output_file, "rb") as attachment:
        # Add file as application/octet-stream
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()

    # send email
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.sendmail(sender_email, receiver_email, text)

    print(f"Successfully sent email to {receiver_email} with {filename} as an attachment")

# ---------------------------------------------
# Main Logic
report_workbook = Workbook()

# Let's create a style template for the header row
header_style = NamedStyle(name="header")
header_style.font = Font(bold=True)
header_style.alignment = Alignment(horizontal="center", vertical="center")

# Utilization Data Operations
utilization_data_worksheet = report_workbook.active
utilization_data_worksheet.title = "Utilization_Data"
utilization_chart_worksheet = report_workbook.create_sheet('Utilization_Chart')

# headers for Utilization Data worksheet
report_headers1 = ['OpCo','Device Name','Interface Name','Interface Id','Interface Speed (Gbps)','Interface Title','Inbound (util %)','Outbound (util %)']
report_headers2 = ['Minimum','Average','Maximum','Minimum','Average','Maximum']
report_rows = 1

# First Utilization Data worksheet header row
for i,header in enumerate(report_headers1, start = 1):
    # Write 'Outbound' in 10th column
    i = 10 if i == 8 else i
    utilization_data_worksheet.cell(row=report_rows , column=i).style = header_style
    utilization_data_worksheet.cell(row=report_rows , column=i).value = header

report_rows += 1

# Second Utilization Data worksheet header row
for i,header in enumerate(report_headers2, start = 7):
    utilization_data_worksheet.cell(row=report_rows , column=i).style = header_style
    utilization_data_worksheet.cell(row=report_rows , column=i).value = header

# Merge cells in Utilization Data worksheet headers
utilization_data_worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
utilization_data_worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
utilization_data_worksheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
utilization_data_worksheet.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
utilization_data_worksheet.merge_cells(start_row=1, start_column=5, end_row=2, end_column=5)
utilization_data_worksheet.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)
utilization_data_worksheet.merge_cells(start_row=1, start_column=7, end_row=1, end_column=9)
utilization_data_worksheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=12)

report_rows += 1

# Utilization Chart Operations
# headers for Utilization Chart worksheet
report_headers3 = ['Device Name','Interface Name','Inbound - Avg (util %)','Inbound - Max (util %)','Outbound - Avg (util %)','Outbound - Max (util %)']
chart_rows = 1

# Utilization Chart worksheet header row
for i, header in enumerate(report_headers3, start=1):
    utilization_chart_worksheet.cell(row=chart_rows , column=i).style = header_style
    utilization_chart_worksheet.cell(row=chart_rows , column=i).value = header

chart_rows += 1

prepare_utilization_data_worksheet()
prepare_utilization_chart_worksheet()
send_email_with_worksheet_attachment()